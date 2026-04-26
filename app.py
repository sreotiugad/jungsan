"""
광고 정산 자동화 시스템 - Flask 백엔드
실행: python app.py
접속: http://localhost:5000
"""

from flask import Flask, request, jsonify, session, send_file, redirect, url_for
from authlib.integrations.flask_client import OAuth
import hashlib, os, io
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__, static_folder='static', static_url_path='')
app.secret_key = os.environ.get('SECRET_KEY', 'settlement_secret_key_change_in_production')
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
app.config['SESSION_COOKIE_SECURE'] = False
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_NAME'] = 'settlement_session'

# Railway 리버스 프록시 설정
from werkzeug.middleware.proxy_fix import ProxyFix
app.wsgi_app = ProxyFix(app.wsgi_app, x_proto=1, x_host=1)

SUPERADMIN_EMAIL = 'taeyang.park@adef.co.kr'

DATABASE_URL = os.environ.get('DATABASE_URL', '')
# Google OAuth 설정
app.config['GOOGLE_CLIENT_ID'] = os.environ.get('GOOGLE_CLIENT_ID', '')
app.config['GOOGLE_CLIENT_SECRET'] = os.environ.get('GOOGLE_CLIENT_SECRET', '')
app.config['GOOGLE_METADATA_URL'] = 'https://accounts.google.com/.well-known/openid-configuration'

oauth = OAuth(app)
google = oauth.register(
    name='google',
    client_id=app.config['GOOGLE_CLIENT_ID'],
    client_secret=app.config['GOOGLE_CLIENT_SECRET'],
    server_metadata_url=app.config['GOOGLE_METADATA_URL'],
    client_kwargs={'scope': 'openid email profile'},
)

# PostgreSQL 또는 SQLite 선택
if DATABASE_URL:
    import psycopg2
    import psycopg2.extras
    def get_db():
        # Railway는 postgres:// 로 주는데 psycopg2는 postgresql:// 필요
        url = DATABASE_URL.replace('postgres://', 'postgresql://', 1)
        conn = psycopg2.connect(url)
        return conn
    PG = True
else:
    import sqlite3
    DB_PATH = 'settlement.db'
    def get_db():
        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row
        return conn
    PG = False

def init_db():
    tables_sql_pg = [
        """CREATE TABLE IF NOT EXISTS users (id SERIAL PRIMARY KEY, username TEXT UNIQUE NOT NULL, password_hash TEXT NOT NULL, name TEXT NOT NULL, team TEXT DEFAULT '', role TEXT DEFAULT 'member')""",
        """CREATE TABLE IF NOT EXISTS advertisers (id SERIAL PRIMARY KEY, user_id INTEGER NOT NULL, name TEXT NOT NULL, biz_no TEXT DEFAULT '', email TEXT DEFAULT '', contact_name TEXT DEFAULT '')""",
        """CREATE TABLE IF NOT EXISTS campaigns (id SERIAL PRIMARY KEY, advertiser_id INTEGER NOT NULL, name TEXT NOT NULL, sort_order INTEGER DEFAULT 0)""",
        """CREATE TABLE IF NOT EXISTS media_rates (id SERIAL PRIMARY KEY, campaign_id INTEGER NOT NULL, media TEXT NOT NULL, markup_rate REAL DEFAULT 0, agency_fee_rate REAL DEFAULT 0, payback_rate REAL DEFAULT 0, sort_order INTEGER DEFAULT 0)""",
        """CREATE TABLE IF NOT EXISTS naver_accounts (id SERIAL PRIMARY KEY, media_rate_id INTEGER NOT NULL, account_no TEXT NOT NULL, account_name TEXT DEFAULT '')""",
        """CREATE TABLE IF NOT EXISTS settlements (id SERIAL PRIMARY KEY, user_id INTEGER NOT NULL, advertiser TEXT, campaign TEXT, media TEXT, period TEXT, start_date TEXT, end_date TEXT, supply_amt REAL DEFAULT 0, markup_rate REAL DEFAULT 0, markup REAL DEFAULT 0, agency_fee_rate REAL DEFAULT 0, agency_fee REAL DEFAULT 0, total REAL DEFAULT 0, billing_date TEXT, prev_diff REAL DEFAULT 0, billing_ad_cost REAL DEFAULT 0, billing_markup REAL DEFAULT 0, billing_total REAL DEFAULT 0, diff REAL DEFAULT 0, account_id TEXT DEFAULT '', note TEXT DEFAULT '', fx_currency TEXT DEFAULT '', fx_rate REAL, created_at TEXT)""",
    ]
    if PG:
        conn = get_db()
        cur = conn.cursor()
        for sql in tables_sql_pg:
            cur.execute(sql)
        conn.commit(); cur.close(); conn.close()
    else:
        import sqlite3 as _sq
        with _sq.connect('settlement.db') as conn:
            for sql in [s.replace('SERIAL', 'INTEGER').replace('PRIMARY KEY)', 'PRIMARY KEY AUTOINCREMENT)') for s in tables_sql_pg]:
                conn.execute(sql)

init_db()


# ───────────── DB 헬퍼 ──────────────────────────────
def db_execute(conn, sql, params=()):
    """SQLite/PostgreSQL 통합 실행. ? → %s 자동 변환"""
    if PG:
        sql = sql.replace('?', '%s')
        cur = conn.cursor()
        cur.execute(sql, params)
        return cur
    else:
        return conn.execute(sql, params)

def db_fetchall(conn, sql, params=()):
    cur = db_execute(conn, sql, params)
    rows = cur.fetchall()
    if PG:
        cols = [d[0] for d in cur.description]
        return [dict(zip(cols, r)) for r in rows]
    else:
        return [dict(r) for r in rows]

def db_fetchone(conn, sql, params=()):
    cur = db_execute(conn, sql, params)
    row = cur.fetchone()
    if row is None: return None
    if PG:
        cols = [d[0] for d in cur.description]
        return dict(zip(cols, row))
    else:
        return dict(row)

def db_insert(conn, sql, params=()):
    """INSERT 후 lastrowid 반환"""
    if PG:
        sql = sql.replace('?', '%s')
        if 'RETURNING id' not in sql:
            sql += ' RETURNING id'
        cur = conn.cursor()
        cur.execute(sql, params)
        return cur.fetchone()[0]
    else:
        cur = conn.execute(sql, params)
        return cur.lastrowid

def db_commit(conn):
    if PG: conn.commit()

# ───────────── 유틸 ───────────────────────────────────
def hash_pw(pw):
    return hashlib.sha256(pw.encode('utf-8')).hexdigest()

def require_login(f):
    from functools import wraps
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return jsonify({'error': '로그인이 필요합니다'}), 401
        return f(*args, **kwargs)
    return decorated


# ───────────── 인증 API ───────────────────────────────
@app.route('/api/register', methods=['POST'])
def register():
    d = request.json
    if not d.get('username') or not d.get('password') or not d.get('name'):
        return jsonify({'error': '필수 항목을 입력해주세요'}), 400
    try:
        with get_db() as conn:
            conn.execute(
                'INSERT INTO users (username, password_hash, name, team) VALUES (?,?,?,?)',
                (d['username'], hash_pw(d['password']), d['name'], d.get('team', ''))
            )
        return jsonify({'success': True})
    except sqlite3.IntegrityError:
        return jsonify({'error': '이미 사용 중인 아이디입니다'}), 400

@app.route('/api/login', methods=['POST'])
def login():
    d = request.json
    with get_db() as conn:
        user = conn.execute(
            'SELECT * FROM users WHERE username=? AND password_hash=?',
            (d.get('username', ''), hash_pw(d.get('password', '')))
        ).fetchone()
    if not user:
        return jsonify({'error': '아이디 또는 비밀번호가 틀렸습니다'}), 401
    session.permanent = True
    session['user_id'] = user['id']
    session['user_name'] = user['name']
    session['user_team'] = user['team']
    return jsonify({'success': True, 'name': user['name'], 'team': user['team']})

@app.route('/api/logout', methods=['POST'])
def logout():
    session.clear()
    return jsonify({'success': True})

@app.route('/api/me')
def me():
    if 'user_id' not in session:
        return jsonify({'error': 'Unauthorized'}), 401
    # DB에서 최신 role 직접 읽기
    try:
        if PG:
            conn = get_db()
            user = db_fetchone(conn, 'SELECT role FROM users WHERE id=?', (session['user_id'],))
            conn.close()
            role = user['role'] if user else 'member'
        else:
            import sqlite3 as _sq
            with _sq.connect('settlement.db') as conn:
                conn.row_factory = _sq.Row
                u = conn.execute('SELECT role FROM users WHERE id=?', (session['user_id'],)).fetchone()
                role = dict(u)['role'] if u else 'member'
    except:
        role = session.get('user_role', 'member')
    return jsonify({
        'id': session['user_id'],
        'name': session['user_name'],
        'team': session.get('user_team', ''),
        'role': role,
        'username': session.get('user_email', session.get('user_id', ''))
    })


# ───────────── Google OAuth ──────────────────────────────
@app.route('/auth/google')
def google_login():
    # Railway/프로덕션에서 https 강제
    redirect_uri = url_for('google_callback', _external=True)
    redirect_uri = redirect_uri.replace('http://', 'https://')
    return google.authorize_redirect(redirect_uri)

@app.route('/auth/google/callback')
def google_callback():
    token = google.authorize_access_token()
    user_info = token.get('userinfo')
    if not user_info:
        return redirect('/?error=google_failed')

    email = user_info.get('email', '')
    name  = user_info.get('name', '')
    role  = 'admin' if email == SUPERADMIN_EMAIL else 'member'

    # DB에서 유저 찾기 또는 생성
    if PG:
        conn = get_db()
        user = db_fetchone(conn, 'SELECT * FROM users WHERE username=?', (email,))
        if not user:
            uid = db_insert(conn, 'INSERT INTO users (username, password_hash, name, team, role) VALUES (?,?,?,?,?)',
                           (email, 'GOOGLE_AUTH', name, '', role))
            db_commit(conn)
            user = db_fetchone(conn, 'SELECT * FROM users WHERE id=?', (uid,))
        else:
            # superadmin이면 role 강제 업데이트
            if email == SUPERADMIN_EMAIL and user.get('role') != 'admin':
                db_execute(conn, 'UPDATE users SET role=? WHERE id=?', ('admin', user['id']))
                db_commit(conn)
                user['role'] = 'admin'
        conn.close()
    else:
        import sqlite3 as _sq
        with _sq.connect('settlement.db') as conn:
            conn.row_factory = _sq.Row
            user = conn.execute('SELECT * FROM users WHERE username=?', (email,)).fetchone()
            if not user:
                conn.execute('INSERT INTO users (username, password_hash, name, team, role) VALUES (?,?,?,?,?)',
                            (email, 'GOOGLE_AUTH', name, '', role))
                conn.commit()
                user = conn.execute('SELECT * FROM users WHERE username=?', (email,)).fetchone()
            user = dict(user)

    session.permanent = True
    session['user_id']    = user['id']
    session['user_name']  = user['name']
    session['user_team']  = user.get('team', '')
    session['user_role']  = user.get('role', 'member')
    session['user_email'] = email
    return redirect('/')

# ───────────── 프로필 API ──────────────────────────────
@app.route('/api/profile')
@require_login
def get_profile():
    if PG:
        conn = get_db()
        user = db_fetchone(conn, 'SELECT * FROM users WHERE id=?', (session['user_id'],))
        conn.close()
    else:
        import sqlite3 as _sq
        with _sq.connect('settlement.db') as conn:
            conn.row_factory = _sq.Row
            user = dict(conn.execute('SELECT * FROM users WHERE id=?', (session['user_id'],)).fetchone())
    return jsonify({
        'id': user['id'], 'name': user['name'],
        'team': user.get('team', ''), 'username': user['username'],
        'role': user.get('role', 'member')
    })

@app.route('/api/profile', methods=['PUT'])
@require_login
def update_profile():
    d = request.json
    name = d.get('name', '').strip()
    team = d.get('team', '').strip()
    if not name:
        return jsonify({'error': '이름을 입력해주세요'}), 400
    if PG:
        conn = get_db()
        db_execute(conn, 'UPDATE users SET name=?, team=? WHERE id=?', (name, team, session['user_id']))
        db_commit(conn)
        conn.close()
    else:
        import sqlite3 as _sq
        with _sq.connect('settlement.db') as conn:
            conn.execute('UPDATE users SET name=?, team=? WHERE id=?', (name, team, session['user_id']))
    session['user_name'] = name
    session['user_team'] = team
    return jsonify({'success': True})

# ───────────── 관리자 API ─────────────────────────────
def require_admin(f):
    from functools import wraps
    @wraps(f)
    def decorated(*args, **kwargs):
        if session.get('user_role') not in ('admin',):
            return jsonify({'error': '관리자 권한이 필요합니다'}), 403
        return f(*args, **kwargs)
    return decorated

def require_manager(f):
    from functools import wraps
    @wraps(f)
    def decorated(*args, **kwargs):
        if session.get('user_role') not in ('admin', 'manager'):
            return jsonify({'error': '팀장 이상 권한이 필요합니다'}), 403
        return f(*args, **kwargs)
    return decorated

@app.route('/api/admin/users')
@require_login
@require_admin
def admin_get_users():
    if PG:
        conn = get_db()
        users = db_fetchall(conn, 'SELECT id, username, name, team, role FROM users ORDER BY role, name')
        conn.close()
    else:
        import sqlite3 as _sq
        with _sq.connect('settlement.db') as conn:
            conn.row_factory = _sq.Row
            users = [dict(r) for r in conn.execute('SELECT id, username, name, team, role FROM users ORDER BY role, name').fetchall()]
    return jsonify(users)

@app.route('/api/admin/users/<int:uid>/role', methods=['PUT'])
@require_login
@require_admin
def admin_set_role(uid):
    role = request.json.get('role', 'member')
    if role not in ('admin', 'manager', 'member'):
        return jsonify({'error': '잘못된 역할입니다'}), 400
    if PG:
        conn = get_db()
        db_execute(conn, 'UPDATE users SET role=? WHERE id=?', (role, uid))
        db_commit(conn)
        conn.close()
    else:
        import sqlite3 as _sq
        with _sq.connect('settlement.db') as conn:
            conn.execute('UPDATE users SET role=? WHERE id=?', (role, uid))
    return jsonify({'success': True})

@app.route('/api/admin/settlements')
@require_login
@require_manager
def admin_get_settlements():
    """팀장/관리자용 - 전체 정산 내역"""
    if PG:
        conn = get_db()
        rows = db_fetchall(conn, """
            SELECT s.*, u.name as user_name, u.team as user_team
            FROM settlements s
            JOIN users u ON s.user_id = u.id
            ORDER BY s.created_at DESC LIMIT 500
        """)
        conn.close()
    else:
        import sqlite3 as _sq
        with _sq.connect('settlement.db') as conn:
            conn.row_factory = _sq.Row
            rows = [dict(r) for r in conn.execute("""
                SELECT s.*, u.name as user_name, u.team as user_team
                FROM settlements s
                JOIN users u ON s.user_id = u.id
                ORDER BY s.created_at DESC LIMIT 500
            """).fetchall()]
    return jsonify(rows)


# ───────────── 광고주 API ─────────────────────────────
def get_adv_full(conn, adv_id):
    adv = conn.execute('SELECT * FROM advertisers WHERE id=?', (adv_id,)).fetchone()
    if not adv:
        return None
    camps = conn.execute('SELECT * FROM campaigns WHERE advertiser_id=? ORDER BY sort_order', (adv_id,)).fetchall()
    campaigns = []
    for c in camps:
        rates = conn.execute('SELECT * FROM media_rates WHERE campaign_id=? ORDER BY sort_order', (c['id'],)).fetchall()
        media_rates_list = []
        for r in rates:
            accts = conn.execute('SELECT * FROM naver_accounts WHERE media_rate_id=?', (r['id'],)).fetchall()
            media_rates_list.append({
                'id': r['id'], 'media': r['media'],
                'markupRate': r['markup_rate'],
                'agencyFeeRate': r['agency_fee_rate'],
                'paybackRate': r['payback_rate'],
                'naverAccounts': [{'id': a['id'], 'accountNo': a['account_no'], 'accountName': a['account_name']} for a in accts]
            })
        campaigns.append({
            'id': c['id'], 'name': c['name'],
            'mediaRates': media_rates_list
        })
    return {
        'id': adv['id'], 'name': adv['name'],
        'bizNo': adv['biz_no'], 'email': adv['email'],
        'contactName': adv['contact_name'], 'campaigns': campaigns
    }

@app.route('/api/advertisers')
@require_login
def get_advertisers():
    with get_db() as conn:
        advs = conn.execute('SELECT id FROM advertisers WHERE user_id=?', (session['user_id'],)).fetchall()
        return jsonify([get_adv_full(conn, a['id']) for a in advs])

@app.route('/api/advertisers', methods=['POST'])
@require_login
def create_advertiser():
    d = request.json
    with get_db() as conn:
        cur = conn.execute(
            'INSERT INTO advertisers (user_id, name, biz_no, email, contact_name) VALUES (?,?,?,?,?)',
            (session['user_id'], d['name'], d.get('bizNo',''), d.get('email',''), d.get('contactName',''))
        )
        adv_id = cur.lastrowid
        for ci, camp in enumerate(d.get('campaigns', [])):
            cur2 = conn.execute('INSERT INTO campaigns (advertiser_id, name, sort_order) VALUES (?,?,?)', (adv_id, camp['name'], ci))
            camp_id = cur2.lastrowid
            for mi, mr in enumerate(camp.get('mediaRates', [])):
                cur3 = conn.execute(
                    'INSERT INTO media_rates (campaign_id, media, markup_rate, agency_fee_rate, payback_rate, sort_order) VALUES (?,?,?,?,?,?)',
                    (camp_id, mr['media'], mr.get('markupRate',0), mr.get('agencyFeeRate',0), mr.get('paybackRate',0), mi)
                )
                mr_id = cur3.lastrowid
                for acct in mr.get('naverAccounts', []):
                    if acct.get('accountNo','').strip():
                        conn.execute('INSERT INTO naver_accounts (media_rate_id, account_no, account_name) VALUES (?,?,?)',
                                     (mr_id, acct['accountNo'].strip(), acct.get('accountName','').strip()))
    return jsonify({'success': True, 'id': adv_id})

@app.route('/api/advertisers/<int:adv_id>', methods=['PUT'])
@require_login
def update_advertiser(adv_id):
    d = request.json
    with get_db() as conn:
        conn.execute(
            'UPDATE advertisers SET name=?, biz_no=?, email=?, contact_name=? WHERE id=? AND user_id=?',
            (d['name'], d.get('bizNo',''), d.get('email',''), d.get('contactName',''), adv_id, session['user_id'])
        )
        camps = conn.execute('SELECT id FROM campaigns WHERE advertiser_id=?', (adv_id,)).fetchall()
        for c in camps:
            conn.execute('DELETE FROM media_rates WHERE campaign_id=?', (c['id'],))
        conn.execute('DELETE FROM campaigns WHERE advertiser_id=?', (adv_id,))
        for ci, camp in enumerate(d.get('campaigns', [])):
            cur = conn.execute('INSERT INTO campaigns (advertiser_id, name, sort_order) VALUES (?,?,?)', (adv_id, camp['name'], ci))
            camp_id = cur.lastrowid
            for mi, mr in enumerate(camp.get('mediaRates', [])):
                cur3 = conn.execute(
                    'INSERT INTO media_rates (campaign_id, media, markup_rate, agency_fee_rate, payback_rate, sort_order) VALUES (?,?,?,?,?,?)',
                    (camp_id, mr['media'], mr.get('markupRate',0), mr.get('agencyFeeRate',0), mr.get('paybackRate',0), mi)
                )
                mr_id = cur3.lastrowid
                for acct in mr.get('naverAccounts', []):
                    if acct.get('accountNo','').strip():
                        conn.execute('INSERT INTO naver_accounts (media_rate_id, account_no, account_name) VALUES (?,?,?)',
                                     (mr_id, acct['accountNo'].strip(), acct.get('accountName','').strip()))
    return jsonify({'success': True})

@app.route('/api/advertisers/<int:adv_id>', methods=['DELETE'])
@require_login
def delete_advertiser(adv_id):
    with get_db() as conn:
        camps = conn.execute('SELECT id FROM campaigns WHERE advertiser_id=?', (adv_id,)).fetchall()
        for c in camps:
            mrs = conn.execute('SELECT id FROM media_rates WHERE campaign_id=?', (c['id'],)).fetchall()
            for mr in mrs:
                conn.execute('DELETE FROM naver_accounts WHERE media_rate_id=?', (mr['id'],))
            conn.execute('DELETE FROM media_rates WHERE campaign_id=?', (c['id'],))
        conn.execute('DELETE FROM campaigns WHERE advertiser_id=?', (adv_id,))
        conn.execute('DELETE FROM advertisers WHERE id=? AND user_id=?', (adv_id, session['user_id']))
    return jsonify({'success': True})


# ───────────── 정산 API ───────────────────────────────
@app.route('/api/settlements')
@require_login
def get_settlements():
    with get_db() as conn:
        rows = conn.execute(
            'SELECT * FROM settlements WHERE user_id=? ORDER BY created_at DESC LIMIT 200',
            (session['user_id'],)
        ).fetchall()
    return jsonify([dict(r) for r in rows])

@app.route('/api/settlements', methods=['POST'])
@require_login
def save_settlements():
    rows = request.json
    now = datetime.now().isoformat()
    with get_db() as conn:
        for r in rows:
            conn.execute('''
                INSERT INTO settlements (user_id, advertiser, campaign, media, period,
                    start_date, end_date, supply_amt, markup_rate, markup,
                    agency_fee_rate, agency_fee, total, billing_date, prev_diff,
                    billing_ad_cost, billing_markup, billing_total, diff,
                    account_id, note, created_at)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            ''', (
                session['user_id'], r.get('advertiser'), r.get('campaign'), r.get('media'),
                r.get('period'), r.get('startDate'), r.get('endDate'),
                r.get('supplyAmt',0), r.get('markupRate',0), r.get('markup',0),
                r.get('agencyFeeRate',0), r.get('agencyFee',0), r.get('total',0),
                r.get('billingDate'), r.get('prevDiff',0), r.get('billingAdCost',0),
                r.get('billingMarkup',0), r.get('billingTotal',0), r.get('diff',0),
                r.get('accountId',''), r.get('note',''), r.get('fxCurrency',''), r.get('fxRate'), now
            ))
    return jsonify({'success': True})


# ───────────── 파일 파싱 API ──────────────────────────
@app.route('/api/parse/naver', methods=['POST'])
@require_login
def parse_naver():
    if 'file' not in request.files:
        return jsonify({'error': '파일이 없습니다'}), 400

    file = request.files['file']
    ext = file.filename.rsplit('.', 1)[-1].lower() if '.' in file.filename else ''
    # 필터링할 계정번호 목록 (프론트에서 전달, 없으면 전체)
    filter_accounts = request.form.getlist('accounts')  # ['1836943', '1937341', ...]
    result = {}   # {account_no: amount}
    all_rows = [] # 전체 파싱 결과 (디버그용)

    try:
        file_bytes = file.read()
        parsed_rows = []  # [{account_no, account_name, media_name, amount}]

        if ext == 'xls':
            import xlrd
            wb = xlrd.open_workbook(file_contents=file_bytes)
            sheet_name = next((n for n in wb.sheet_names() if '세금' in n or '계산서' in n), None)
            sheet = wb.sheet_by_name(sheet_name) if sheet_name else wb.sheet_by_index(0)

            hdr_row, cols = -1, {}
            for i in range(min(sheet.nrows, 15)):
                vals = [str(sheet.cell_value(i, j)).strip() for j in range(sheet.ncols)]
                if any('계정명' in v for v in vals) or any('공급가액' in v for v in vals):
                    hdr_row = i
                    for j, v in enumerate(vals):
                        cols[v] = j
                    break

            if hdr_row >= 0:
                acct_col = next((j for v, j in cols.items() if '계정명' in v), -1)
                amt_col  = next((j for v, j in cols.items() if '공급가액' in v), -1)
                prod_col = next((j for v, j in cols.items() if '품목' in v or '상품' in v), -1)
                note_col = next((j for v, j in cols.items() if '비고' in v), -1)

                for i in range(hdr_row + 1, sheet.nrows):
                    if acct_col < 0 or amt_col < 0:
                        break
                    acct_name = str(sheet.cell_value(i, acct_col)).strip()
                    note_val  = str(sheet.cell_value(i, note_col)).strip() if note_col >= 0 else ''
                    prod_val  = str(sheet.cell_value(i, prod_col)).strip() if prod_col >= 0 else ''
                    try:
                        amt = sheet.cell_value(i, amt_col)
                        if isinstance(amt, (int, float)) and amt > 0 and acct_name not in ('', 'None'):
                            # 비고에서 계정번호 추출 (숫자만 있는 경우)
                            account_no = ''
                            try:
                                clean = note_val.replace(',', '').strip()
                                if clean.isdigit():
                                    account_no = clean
                            except:
                                pass
                            parsed_rows.append({
                                'accountNo': account_no,
                                'accountName': acct_name,
                                'media': prod_val,
                                'amount': float(amt)
                            })
                    except:
                        pass

        else:  # xlsx
            import io as _io
            wb = openpyxl.load_workbook(_io.BytesIO(file_bytes), read_only=True, data_only=True)
            sheet_name = next((n for n in wb.sheetnames if '세금' in n or '계산서' in n), None)
            ws = wb[sheet_name] if sheet_name else wb.active

            rows_data = list(ws.iter_rows(values_only=True))
            hdr_row, cols = -1, {}
            for i, row in enumerate(rows_data[:15]):
                if not row:
                    continue
                strs = [str(c).strip() if c is not None else '' for c in row]
                if any('계정명' in v for v in strs) or any('공급가액' in v for v in strs):
                    hdr_row = i
                    cols = {v: j for j, v in enumerate(strs)}
                    break

            if hdr_row >= 0:
                acct_col = next((j for v, j in cols.items() if '계정명' in v), -1)
                amt_col  = next((j for v, j in cols.items() if '공급가액' in v), -1)
                prod_col = next((j for v, j in cols.items() if '품목' in v or '상품' in v), -1)
                note_col = next((j for v, j in cols.items() if '비고' in v), -1)

                for row in rows_data[hdr_row + 1:]:
                    if not row or acct_col < 0 or amt_col < 0:
                        continue
                    acct_name = str(row[acct_col]).strip() if row[acct_col] is not None else ''
                    note_val  = str(row[note_col]).strip() if note_col >= 0 and row[note_col] is not None else ''
                    prod_val  = str(row[prod_col]).strip() if prod_col >= 0 and row[prod_col] is not None else ''
                    amt = row[amt_col]
                    if isinstance(amt, (int, float)) and amt > 0 and acct_name not in ('None', ''):
                        account_no = ''
                        try:
                            clean = note_val.replace(',', '').strip()
                            if clean.isdigit():
                                account_no = clean
                        except:
                            pass
                        parsed_rows.append({
                            'accountNo': account_no,
                            'accountName': acct_name,
                            'media': prod_val,
                            'amount': float(amt)
                        })

    except Exception as e:
        return jsonify({'error': f'파싱 오류: {str(e)}'}), 500

    if not parsed_rows:
        return jsonify({'error': '금액 데이터를 찾지 못했습니다. 네이버 위임발행 파일이 맞는지 확인해주세요'}), 400

    # 계정번호 필터링 적용
    if filter_accounts:
        filtered = [r for r in parsed_rows if r['accountNo'] in filter_accounts]
    else:
        filtered = parsed_rows

    # 계정번호별 합계
    by_account = {}
    for r in filtered:
        key = r['accountNo'] or r['accountName']
        if key not in by_account:
            by_account[key] = {'accountNo': r['accountNo'], 'accountName': r['accountName'], 'total': 0, 'rows': []}
        by_account[key]['total'] += r['amount']
        by_account[key]['rows'].append({'media': r['media'], 'amount': r['amount']})

    total = sum(v['total'] for v in by_account.values())

    return jsonify({
        'data': by_account,
        'total': total,
        'allAccounts': list({r['accountNo']: r['accountName'] for r in parsed_rows if r['accountNo']}.items())
    })



# ───────────── 메타 PDF 파싱 API ──────────────────────
@app.route('/api/parse/meta', methods=['POST'])
@require_login
def parse_meta():
    if 'file' not in request.files:
        return jsonify({'error': '파일이 없습니다'}), 400

    file = request.files['file']
    filter_accounts = request.form.getlist('accounts')  # 등록된 Account Id 목록
    ext = file.filename.rsplit('.', 1)[-1].lower() if '.' in file.filename else ''
    results = []

    try:
        import zipfile, io as _io, pdfplumber, re

        file_bytes = file.read()

        def parse_single_meta_pdf(pdf_bytes, filename=''):
            try:
                with pdfplumber.open(_io.BytesIO(pdf_bytes)) as pdf:
                    text = '\n'.join(p.extract_text() or '' for p in pdf.pages)

                invoice_no  = ''
                account_id  = ''
                advertiser  = ''
                subtotal    = 0
                period_str  = ''
                currency    = 'KRW'

                # Invoice # 추출
                m = re.search(r'Invoice\s*#[:\s]+([\d]+)', text)
                if m:
                    invoice_no = m.group(1).strip()

                # Account Id / Group 추출
                m2 = re.search(r'Account\s*Id\s*/\s*Group[:\s]+([\d]+)', text)
                if m2:
                    account_id = m2.group(1).strip()

                # Advertiser 추출
                m3 = re.search(r'Advertiser[:\s]+(.+?)\n', text)
                if m3:
                    advertiser = m3.group(1).strip()

                # Billing Period 추출
                m4 = re.search(r'Billing\s*Period[:\s]+([A-Za-z]+-\d+)', text)
                if m4:
                    period_str = m4.group(1).strip()

                # Subtotal 추출 (VAT 제외 금액)
                m5 = re.search(r'Subtotal[:\s]+([\d,]+)', text)
                if m5:
                    subtotal = int(m5.group(1).replace(',', ''))

                # Invoice Currency 추출
                m6 = re.search(r'Invoice\s*Currency[:\s]+([A-Z]{3})', text)
                if m6:
                    currency = m6.group(1).strip()

                if (invoice_no or account_id) and subtotal > 0:
                    return {
                        'invoiceNo':  invoice_no,
                        'accountId':  account_id,
                        'advertiser': advertiser,
                        'period':     period_str,
                        'subtotal':   subtotal,
                        'currency':   currency,
                        'filename':   filename,
                    }
            except Exception as e:
                pass
            return None

        if ext == 'zip':
            with zipfile.ZipFile(_io.BytesIO(file_bytes)) as zf:
                pdf_files = [n for n in zf.namelist() if n.lower().endswith('.pdf')]
                for pdf_name in pdf_files:
                    with zf.open(pdf_name) as pf:
                        r = parse_single_meta_pdf(pf.read(), pdf_name)
                        if r:
                            results.append(r)
        elif ext == 'pdf':
            r = parse_single_meta_pdf(file_bytes, file.filename)
            if r:
                results.append(r)
        else:
            return jsonify({'error': 'ZIP 또는 PDF 파일만 지원합니다'}), 400

    except Exception as e:
        return jsonify({'error': f'파싱 오류: {str(e)}'}), 500

    if not results:
        return jsonify({'error': '메타 인보이스에서 데이터를 찾지 못했습니다'}), 400

    # Account Id 기준 필터링
    if filter_accounts:
        filtered = [r for r in results if r['accountId'] in filter_accounts]
    else:
        filtered = results

    total = sum(r['subtotal'] for r in filtered)

    return jsonify({
        'data':        {r['accountId'] or r['invoiceNo']: r for r in filtered},
        'total':       total,
        'allAccounts': results,
    })

# ───────────── 트위터(X) PDF 파싱 API ─────────────────
@app.route('/api/parse/twitter', methods=['POST'])
@require_login
def parse_twitter():
    if 'file' not in request.files:
        return jsonify({'error': '파일이 없습니다'}), 400

    file = request.files['file']
    filter_accounts = request.form.getlist('accounts')
    ext = file.filename.rsplit('.', 1)[-1].lower() if '.' in file.filename else ''
    results = []

    try:
        import zipfile, io as _io, pdfplumber, re
        file_bytes = file.read()

        def parse_single_twitter_pdf(pdf_bytes, filename=''):
            try:
                with pdfplumber.open(_io.BytesIO(pdf_bytes)) as pdf:
                    text = '\n'.join(p.extract_text() or '' for p in pdf.pages)

                invoice_no = ''
                account_no = ''
                advertiser = ''
                period_str = ''
                subtotal   = 0
                currency   = 'KRW'

                # Invoice Number
                m = re.search(r'Invoice\s*Number\s*[\n\r]+([\d]+)', text)
                if m:
                    invoice_no = m.group(1).strip()

                # Account No.
                m2 = re.search(r'Account\s*No\.?\s*[:\s]+([\d]+)', text)
                if m2:
                    account_no = m2.group(1).strip()

                # Advertiser
                m3 = re.search(r'Advertiser[:\s]+(.+?)\n', text)
                if m3:
                    advertiser = m3.group(1).strip()

                # Invoice Period
                m4 = re.search(r'Invoice\s*Period\s*[\n\r]+([A-Z]+-\d+)', text)
                if m4:
                    period_str = m4.group(1).strip()

                # Sub-Total (VAT 제외)
                m5 = re.search(r'Sub-Total\s+[A-Z]{3}\s+([\d,]+\.?\d*)', text)
                if m5:
                    subtotal = int(float(m5.group(1).replace(',', '')))

                # Currency
                m6 = re.search(r'Sub-Total\s+([A-Z]{3})', text)
                if m6:
                    currency = m6.group(1).strip()

                if (invoice_no or account_no) and subtotal > 0:
                    return {
                        'invoiceNo':  invoice_no,
                        'accountId':  account_no,
                        'advertiser': advertiser,
                        'period':     period_str,
                        'subtotal':   subtotal,
                        'currency':   currency,
                        'filename':   filename,
                    }
            except Exception as e:
                pass
            return None

        if ext == 'zip':
            with zipfile.ZipFile(_io.BytesIO(file_bytes)) as zf:
                pdf_files = [n for n in zf.namelist() if n.lower().endswith('.pdf')]
                for pdf_name in pdf_files:
                    with zf.open(pdf_name) as pf:
                        r = parse_single_twitter_pdf(pf.read(), pdf_name)
                        if r:
                            results.append(r)
        elif ext == 'pdf':
            r = parse_single_twitter_pdf(file_bytes, file.filename)
            if r:
                results.append(r)
        else:
            return jsonify({'error': 'ZIP 또는 PDF 파일만 지원합니다'}), 400

    except Exception as e:
        return jsonify({'error': f'파싱 오류: {str(e)}'}), 500

    if not results:
        return jsonify({'error': '트위터(X) 인보이스에서 데이터를 찾지 못했습니다'}), 400

    # Account No. 기준 필터링
    if filter_accounts:
        filtered = [r for r in results if r['accountId'] in filter_accounts]
    else:
        filtered = results

    total = sum(r['subtotal'] for r in filtered)

    return jsonify({
        'data':        {r['accountId'] or r['invoiceNo']: r for r in filtered},
        'total':       total,
        'allAccounts': results,
    })

# ───────────── 네이버 GFA 파싱 API ───────────────────
@app.route('/api/parse/naver-gfa', methods=['POST'])
@require_login
def parse_naver_gfa():
    if 'file' not in request.files:
        return jsonify({'error': '파일이 없습니다'}), 400

    file = request.files['file']
    filter_accounts = request.form.getlist('accounts')  # 비고 번호 목록
    ext = file.filename.rsplit('.', 1)[-1].lower() if '.' in file.filename else ''

    try:
        import io as _io
        file_bytes = file.read()

        if ext == 'xls':
            import xlrd
            wb = xlrd.open_workbook(file_contents=file_bytes)
            sheet = wb.sheet_by_index(0)
            rows_data = [
                [sheet.cell_value(r, c) for c in range(sheet.ncols)]
                for r in range(sheet.nrows)
            ]
        else:
            wb = openpyxl.load_workbook(_io.BytesIO(file_bytes), read_only=True, data_only=True)
            ws = wb.active
            rows_data = [list(row) for row in ws.iter_rows(values_only=True)]

        # 헤더 행 찾기
        hdr_row, cols = -1, {}
        for i, row in enumerate(rows_data[:5]):
            strs = [str(c).strip() if c is not None else '' for c in row]
            if any('공급가액' in v for v in strs) and any('비고' in v for v in strs):
                hdr_row = i
                cols = {v: j for j, v in enumerate(strs)}
                break

        if hdr_row < 0:
            return jsonify({'error': '헤더를 찾을 수 없습니다. GFA 위임발행 파일인지 확인해주세요'}), 400

        # 컬럼 인덱스
        amt_col    = next((j for v, j in cols.items() if '공급가액' in v), -1)
        note_col   = next((j for v, j in cols.items() if '비고' in v), -1)
        acct_col   = next((j for v, j in cols.items() if '계정' in v and '이름' in v), -1)
        adv_col    = next((j for v, j in cols.items() if '광고주명' in v), -1)

        # 데이터 파싱
        parsed = {}  # {account_no: {accountNo, accountName, total}}
        for row in rows_data[hdr_row + 1:]:
            if not row or amt_col < 0 or note_col < 0:
                continue
            amt  = row[amt_col]
            note = row[note_col]
            acct_name = str(row[acct_col]).strip() if acct_col >= 0 and row[acct_col] else ''
            adv_name  = str(row[adv_col]).strip() if adv_col >= 0 and row[adv_col] else ''

            if not isinstance(amt, (int, float)) or amt == 0:
                continue

            # 비고에서 계정번호 추출
            account_no = ''
            try:
                clean = str(note).replace(',', '').strip()
                if clean.replace('.', '').isdigit():
                    account_no = str(int(float(clean)))
            except:
                pass

            if not account_no:
                continue

            if account_no not in parsed:
                parsed[account_no] = {
                    'accountNo':   account_no,
                    'accountName': acct_name or adv_name,
                    'total': 0,
                    'rows': []
                }
            parsed[account_no]['total'] += float(amt)
            parsed[account_no]['rows'].append({
                'accountName': acct_name,
                'advertiser':  adv_name,
                'amount':      float(amt)
            })

    except Exception as e:
        return jsonify({'error': f'파싱 오류: {str(e)}'}), 500

    if not parsed:
        return jsonify({'error': '데이터를 찾지 못했습니다'}), 400

    # 필터링
    if filter_accounts:
        filtered = {k: v for k, v in parsed.items() if k in filter_accounts}
    else:
        filtered = parsed

    if not filtered:
        return jsonify({
            'error': f'등록된 계정번호와 일치하는 데이터가 없습니다. 파일 내 계정번호: {", ".join(sorted(parsed.keys())[:10])}',
        }), 400

    total = sum(v['total'] for v in filtered.values())

    return jsonify({
        'data':        filtered,
        'total':       total,
        'allAccounts': [{'accountNo': k, 'accountName': v['accountName'], 'total': v['total']}
                        for k, v in parsed.items()]
    })

# ───────────── 크리테오 파싱 API ──────────────────────
@app.route('/api/parse/criteo', methods=['POST'])
@require_login
def parse_criteo():
    if 'file' not in request.files:
        return jsonify({'error': '파일이 없습니다'}), 400

    file = request.files['file']
    filter_accounts = request.form.getlist('accounts')  # Account Name 목록
    ext = file.filename.rsplit('.', 1)[-1].lower() if '.' in file.filename else ''

    try:
        import io as _io
        file_bytes = file.read()

        if ext == 'xls':
            import xlrd
            wb = xlrd.open_workbook(file_contents=file_bytes)
            ws = wb.sheet_by_index(0)
            rows_data = [[ws.cell_value(r, c) for c in range(ws.ncols)] for r in range(ws.nrows)]
        else:
            wb = openpyxl.load_workbook(_io.BytesIO(file_bytes), read_only=True, data_only=True)
            ws = wb.active
            rows_data = [list(row) for row in ws.iter_rows(values_only=True)]

        # 헤더 행 찾기
        hdr_row, cols = -1, {}
        for i, row in enumerate(rows_data[:5]):
            strs = [str(c).strip() if c is not None else '' for c in row]
            if any('Account Name' in v or 'account' in v.lower() for v in strs):
                hdr_row = i
                cols = {v: j for j, v in enumerate(strs)}
                break

        if hdr_row < 0:
            return jsonify({'error': '헤더를 찾을 수 없습니다. 크리테오 정산파일인지 확인해주세요'}), 400

        # 컬럼 인덱스
        acct_col = next((j for v, j in cols.items() if 'Account Name' in v or 'account name' in v.lower()), -1)
        amt_col  = next((j for v, j in cols.items() if 'Invoicing' in v and 'Amount' in v and 'Total' not in v), -1)

        if acct_col < 0 or amt_col < 0:
            return jsonify({'error': f'Account Name 또는 Invoicing Amount 컬럼을 찾을 수 없습니다'}), 400

        # 데이터 파싱
        parsed = {}
        for row in rows_data[hdr_row + 1:]:
            if not row:
                continue
            acct = str(row[acct_col]).strip() if row[acct_col] is not None else ''
            amt  = row[amt_col]
            if not acct or acct in ('NaN', 'nan', '', 'None') or not isinstance(amt, (int, float)) or amt == 0:
                continue
            if acct not in parsed:
                parsed[acct] = {'accountNo': acct, 'accountName': acct, 'total': 0}
            parsed[acct]['total'] += float(amt)

    except Exception as e:
        return jsonify({'error': f'파싱 오류: {str(e)}'}), 500

    if not parsed:
        return jsonify({'error': '데이터를 찾지 못했습니다'}), 400

    # Account Name으로 필터링 (대소문자 무시)
    if filter_accounts:
        filter_lower = [f.lower().strip() for f in filter_accounts]
        filtered = {k: v for k, v in parsed.items() if k.lower().strip() in filter_lower}
    else:
        filtered = parsed

    if not filtered:
        return jsonify({
            'error': f'등록된 계정명과 일치하는 데이터가 없습니다. 파일 내 계정: {", ".join(list(parsed.keys())[:5])}...'
        }), 400

    total = sum(v['total'] for v in filtered.values())

    return jsonify({
        'data':        filtered,
        'total':       total,
        'allAccounts': [{'accountNo': k, 'accountName': v['accountName'], 'total': v['total']}
                        for k, v in parsed.items()]
    })

# ───────────── 카카오 파싱 API (브검/모먼트/톡채널/클릭스 통합) ──
@app.route('/api/parse/kakao', methods=['POST'])
@require_login
def parse_kakao():
    if 'file' not in request.files:
        return jsonify({'error': '파일이 없습니다'}), 400

    file = request.files['file']
    filter_accounts = request.form.getlist('accounts')  # 자산ID 또는 월렛ID
    ext = file.filename.rsplit('.', 1)[-1].lower() if '.' in file.filename else ''

    try:
        import io as _io
        file_bytes = file.read()

        if ext == 'xls':
            import xlrd
            raw_wb = xlrd.open_workbook(file_contents=file_bytes)
            sheet_names = raw_wb.sheet_names()
        else:
            raw_wb = openpyxl.load_workbook(_io.BytesIO(file_bytes), read_only=True, data_only=True)
            sheet_names = raw_wb.sheetnames

        # 계산서 위임 시트 찾기
        target_sheet = next(
            (s for s in sheet_names if '계산서' in s and ('위임' in s or '발행' in s)
             and '수정' not in s),
            None
        )
        if not target_sheet:
            return jsonify({'error': '계산서 위임 시트를 찾을 수 없습니다'}), 400

        if ext == 'xls':
            ws = raw_wb.sheet_by_name(target_sheet)
            rows_data = [[ws.cell_value(r, c) for c in range(ws.ncols)] for r in range(ws.nrows)]
        else:
            ws = raw_wb[target_sheet]
            rows_data = [list(row) for row in ws.iter_rows(values_only=True)]

        # 헤더 찾기
        hdr_row, cols = -1, {}
        for i, row in enumerate(rows_data[:5]):
            strs = [str(c).strip() if c is not None else '' for c in row]
            if any('공급가액' in v for v in strs):
                hdr_row = i
                cols = {v: j for j, v in enumerate(strs)}
                break

        if hdr_row < 0:
            return jsonify({'error': '헤더(공급가액)를 찾을 수 없습니다'}), 400

        # 컬럼 인덱스
        amt_col   = next((j for v, j in cols.items() if '공급가액' in v), -1)
        name_col  = next((j for v, j in cols.items() if '광고계정명' in v or '자산 이름' in v or '자산이름' in v), -1)
        asset_col = next((j for v, j in cols.items() if '자산' in v and 'ID' in v.upper() or '자산ID' in v), -1)
        wallet_col= next((j for v, j in cols.items() if '월렛' in v and ('ID' in v.upper() or 'ID' in v) and '이름' not in v), -1)

        if amt_col < 0:
            return jsonify({'error': '공급가액 컬럼을 찾을 수 없습니다'}), 400

        # 데이터 파싱
        parsed = {}
        for row in rows_data[hdr_row + 1:]:
            if not row:
                continue
            amt = row[amt_col]
            if not isinstance(amt, (int, float)) or amt == 0:
                continue

            # ID 추출 (자산ID 우선, 없으면 월렛ID)
            acct_id = ''
            if asset_col >= 0 and row[asset_col] is not None:
                v = str(row[asset_col]).strip()
                if v not in ('', 'None', 'nan'):
                    acct_id = v.split('.')[0]  # 소수점 제거
            if not acct_id and wallet_col >= 0 and row[wallet_col] is not None:
                v = str(row[wallet_col]).strip()
                if v not in ('', 'None', 'nan'):
                    acct_id = v.split('.')[0]

            if not acct_id:
                continue

            # 정수형으로 정규화
            try:
                acct_id = str(int(float(acct_id)))
            except:
                pass

            acct_name = str(row[name_col]).strip() if name_col >= 0 and row[name_col] is not None else acct_id

            if acct_id not in parsed:
                parsed[acct_id] = {'accountNo': acct_id, 'accountName': acct_name, 'total': 0}
            parsed[acct_id]['total'] += float(amt)

    except Exception as e:
        return jsonify({'error': f'파싱 오류: {str(e)}'}), 500

    if not parsed:
        return jsonify({'error': '데이터를 찾지 못했습니다'}), 400

    # 필터링 (자산ID/월렛ID)
    if filter_accounts:
        norm = [str(int(float(a))) if a.replace('.','').isdigit() else a for a in filter_accounts]
        filtered = {k: v for k, v in parsed.items() if k in norm}
    else:
        filtered = parsed

    if not filtered:
        return jsonify({
            'error': f'등록된 ID와 일치하는 데이터가 없습니다. 파일 내 ID: {", ".join(list(parsed.keys())[:8])}'
        }), 400

    total = sum(v['total'] for v in filtered.values())
    return jsonify({
        'data':        filtered,
        'total':       total,
        'allAccounts': [{'accountNo': k, 'accountName': v['accountName'], 'total': v['total']}
                        for k, v in parsed.items()]
    })

# ───────────── 와이즈버즈 파싱 API ───────────────────
@app.route('/api/parse/wisebirds', methods=['POST'])
@require_login
def parse_wisebirds():
    if 'file' not in request.files:
        return jsonify({'error': '파일이 없습니다'}), 400

    file = request.files['file']
    # 요청한 매체 목록 (예: ["Meta", "TikTok"])
    target_media = request.form.getlist('media')
    # 메타 광고계정 ID 목록
    meta_accounts = request.form.getlist('metaAccounts')
    ext = file.filename.rsplit('.', 1)[-1].lower() if '.' in file.filename else ''

    try:
        import io as _io, re
        file_bytes = file.read()

        if ext == 'xls':
            import xlrd
            raw_wb = xlrd.open_workbook(file_contents=file_bytes)
            sheet_names = raw_wb.sheet_names()
        else:
            raw_wb = openpyxl.load_workbook(_io.BytesIO(file_bytes), read_only=True, data_only=True)
            sheet_names = raw_wb.sheetnames

        # 가장 최신 시트 (첫 번째 시트)
        sheet_name = sheet_names[0]

        if ext == 'xls':
            ws = raw_wb.sheet_by_name(sheet_name)
            rows_data = [[ws.cell_value(r, c) for c in range(ws.ncols)] for r in range(ws.nrows)]
        else:
            ws = raw_wb[sheet_name]
            rows_data = [list(row) for row in ws.iter_rows(values_only=True)]

        # 매체명 정규화 매핑
        MEDIA_MAP = {
            'Meta': ['Meta', 'meta', '메타'],
            'X(Twitter)': ['X(Twitter)', 'Twitter', 'twitter', 'X', '트위터'],
            'TikTok': ['TikTok', 'tiktok', '틱톡'],
            'Toss': ['Toss', 'toss', '토스'],
            '당근': ['당근', 'Daangn', 'daangn'],
            'Apple': ['Apple', 'apple', '애플'],
        }

        # ── 1. 상단 요약표 파싱 (인보이스 금액 컬럼) ──
        summary = {}  # {표준매체명: 인보이스금액}
        media_col, inv_col = -1, -1

        for i, row in enumerate(rows_data[:15]):
            strs = [str(c).strip() if c is not None else '' for c in row]
            if any('인보이스 금액' in v or '인보이스금액' in v for v in strs):
                media_col = next((j for j, v in enumerate(strs) if '매체' in v), 1)
                inv_col   = next((j for j, v in enumerate(strs) if '인보이스 금액' in v or '인보이스금액' in v), 3)
                # 이후 행들에서 매체별 금액 읽기
                for row2 in rows_data[i+1:]:
                    if not row2: continue
                    media_val = str(row2[media_col]).strip() if row2[media_col] is not None else ''
                    inv_val   = row2[inv_col] if inv_col < len(row2) else None
                    if 'Total' in media_val or not media_val or media_val in ('None', 'nan'): continue
                    if not isinstance(inv_val, (int, float)) or inv_val == 0: continue
                    # 표준 매체명으로 변환
                    for std, aliases in MEDIA_MAP.items():
                        if any(a.lower() in media_val.lower() for a in aliases):
                            summary[std] = float(inv_val)
                            break
                break

        # ── 2. 메타 상세내역 파싱 (광고계정 ID 기반) ──
        meta_detail = {}  # {account_id: {invoiceNo, advertiser, amount}}
        in_meta_detail = False
        acct_id_col = inv_amount_col = adv_col = inv_no_col = -1

        for i, row in enumerate(rows_data):
            strs = [str(c).strip() if c is not None else '' for c in row]
            # 메타 상세내역 헤더 찾기
            if any('인보이스 번호' in v for v in strs) and any('광고 계정 ID' in v or '광고계정ID' in v for v in strs):
                in_meta_detail = True
                inv_no_col    = next((j for j,v in enumerate(strs) if '인보이스 번호' in v), -1)
                adv_col       = next((j for j,v in enumerate(strs) if '광고주' in v), -1)
                inv_amount_col= next((j for j,v in enumerate(strs) if '인보이스 금액' in v or '금액' in v), -1)
                acct_id_col   = next((j for j,v in enumerate(strs) if '광고 계정 ID' in v or '광고계정ID' in v), -1)
                continue
            if in_meta_detail and acct_id_col >= 0:
                if not any(c is not None for c in row): continue
                acct_id = str(row[acct_id_col]).strip() if row[acct_id_col] is not None else ''
                amt     = row[inv_amount_col] if inv_amount_col >= 0 and inv_amount_col < len(row) else None
                inv_no  = str(row[inv_no_col]).strip() if inv_no_col >= 0 and row[inv_no_col] is not None else ''
                adv     = str(row[adv_col]).strip() if adv_col >= 0 and row[adv_col] is not None else ''
                if not acct_id or acct_id in ('None', 'nan', '') or not isinstance(amt, (int, float)) or amt == 0:
                    continue
                if acct_id not in meta_detail:
                    meta_detail[acct_id] = {'accountId': acct_id, 'advertiser': adv, 'invoiceNo': inv_no, 'total': 0}
                meta_detail[acct_id]['total'] += float(amt)

        # ── 3. 결과 조합 ──
        result = {}

        for media in target_media:
            # 표준 매체명 찾기
            std_media = next((k for k, aliases in MEDIA_MAP.items() if any(a.lower() in media.lower() for a in aliases + [media])), media)

            if std_media == 'Meta' and meta_accounts:
                # 메타: 광고계정 ID로 필터링
                filtered = {k: v for k, v in meta_detail.items()
                           if any(acc in k for acc in meta_accounts)}
                total = sum(v['total'] for v in filtered.values())
                result[media] = {
                    'media': media,
                    'total': total,
                    'detail': filtered,
                    'invoiceNos': list({v['invoiceNo'] for v in filtered.values() if v['invoiceNo']}),
                }
            else:
                # 기타 매체: 요약표에서 합계
                total = summary.get(std_media, 0)
                result[media] = {
                    'media': media,
                    'total': total,
                    'detail': {},
                    'invoiceNos': [],
                }

        return jsonify({
            'result':       result,
            'summary':      summary,
            'allMetaAccts': list(meta_detail.keys()),
            'sheetName':    sheet_name,
        })

    except Exception as e:
        import traceback
        return jsonify({'error': f'파싱 오류: {str(e)}', 'trace': traceback.format_exc()}), 500

# ───────────── 구글 ZIP 파싱 API ──────────────────────
@app.route('/api/parse/google', methods=['POST'])
@require_login
def parse_google():
    if 'file' not in request.files:
        return jsonify({'error': '파일이 없습니다'}), 400

    file = request.files['file']
    filter_accounts = request.form.getlist('accounts')  # ['196-947-7609', ...]
    ext = file.filename.rsplit('.', 1)[-1].lower() if '.' in file.filename else ''

    results = []  # [{accountId, accountName, subtotal}]

    try:
        import zipfile, io as _io, pdfplumber, re

        file_bytes = file.read()

        def parse_single_pdf(pdf_bytes, filename=''):
            """PDF 한 개에서 계정ID + 소계 추출"""
            try:
                with pdfplumber.open(_io.BytesIO(pdf_bytes)) as pdf:
                    text = '\n'.join(p.extract_text() or '' for p in pdf.pages)

                account_id   = ''
                account_name = ''
                subtotal     = 0

                # 계정 ID 추출: "계정 ID: 196-947-7609"
                m = re.search(r'계정\s*ID\s*[:\:]?\s*([\d\-]+)', text)
                if m:
                    account_id = m.group(1).strip()

                # 계정명 추출: "계정: ㈜다우기술"
                m2 = re.search(r'계정\s*[:\：]\s*(.+?)\n', text)
                if m2:
                    account_name = m2.group(1).strip()

                # 소계(KRW) 추출 - 첫 번째 매칭만 사용
                m3 = re.search(r'소계\(KRW\)\s*[₩￦]?\s*([\d,]+)', text)
                if m3:
                    subtotal = int(m3.group(1).replace(',', ''))

                if account_id and subtotal > 0:
                    return {'accountId': account_id, 'accountName': account_name,
                            'subtotal': subtotal, 'filename': filename}
            except Exception as e:
                pass
            return None

        if ext == 'zip':
            with zipfile.ZipFile(_io.BytesIO(file_bytes)) as zf:
                pdf_files = [n for n in zf.namelist() if n.lower().endswith('.pdf')]
                for pdf_name in pdf_files:
                    with zf.open(pdf_name) as pf:
                        r = parse_single_pdf(pf.read(), pdf_name)
                        if r:
                            results.append(r)

        elif ext == 'pdf':
            r = parse_single_pdf(file_bytes, file.filename)
            if r:
                results.append(r)

        else:
            return jsonify({'error': 'ZIP 또는 PDF 파일만 지원합니다'}), 400

    except Exception as e:
        return jsonify({'error': f'파싱 오류: {str(e)}'}), 500

    if not results:
        return jsonify({'error': 'PDF에서 계정 정보를 찾지 못했습니다'}), 400

    # 계정번호 필터링
    if filter_accounts:
        filtered = [r for r in results if r['accountId'] in filter_accounts]
    else:
        filtered = results

    total = sum(r['subtotal'] for r in filtered)

    return jsonify({
        'data': {r['accountId']: r for r in filtered},
        'total': total,
        'allAccounts': [{'accountId': r['accountId'], 'accountName': r['accountName'],
                         'subtotal': r['subtotal'], 'filename': r['filename']} for r in results]
    })

# ───────────── 기존 파일 채우기 API ──────────────────
@app.route('/api/export/fill-template', methods=['POST'])
@require_login
def fill_template():
    """기존 세금계산서발행요청파일에 이번달 데이터를 이어 붙여서 반환"""
    import json, io as _io

    if 'template' not in request.files:
        return jsonify({'error': '템플릿 파일이 없습니다'}), 400

    rows       = json.loads(request.form.get('rows', '[]'))
    adv_name   = request.form.get('advName', '')
    camp_name  = request.form.get('campName', '')
    period     = request.form.get('period', '')
    billing_date = request.form.get('billingDate', '')
    adv_biz_no = request.form.get('advBizNo', '')
    adv_email  = request.form.get('advEmail', '')
    user_name  = session.get('user_name', '')
    user_team  = session.get('user_team', '')

    template_file = request.files['template']

    try:
        wb = openpyxl.load_workbook(_io.BytesIO(template_file.read()))
    except Exception as e:
        return jsonify({'error': f'파일을 열 수 없습니다: {str(e)}'}), 400

    # ── 세일즈양식 시트 처리 ──
    if '세일즈양식' not in wb.sheetnames:
        return jsonify({'error': '세일즈양식 시트를 찾을 수 없습니다'}), 400

    ws = wb['세일즈양식']

    # 병합행 목록 미리 수집
    from openpyxl.cell.cell import MergedCell
    merged_rows = set()
    for merge in ws.merged_cells.ranges:
        for rr in range(merge.min_row, merge.max_row + 1):
            merged_rows.add(rr)

    # 마지막 번호(B열) + 실제 데이터 있는 마지막 행 둘 다 찾기
    last_no = 0
    last_data_row = 2
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
        rnum = row[0].row
        if rnum in merged_rows:
            continue
        b = row[1].value   # B열 = 번호
        e = row[4].value   # E열 = 광고주명
        g = row[6].value   # G열 = 매체
        i = row[8].value   # I열 = 매입월
        l = row[11].value  # L열 = 공급가액
        # 실제 데이터가 있는 행이면 last_data_row 갱신
        if any(v is not None for v in [e, g, i, l]):
            last_data_row = rnum
        # 번호 최대값 갱신
        if b is not None:
            try:
                no = int(str(b).strip())
                if no > last_no:
                    last_no = no
            except:
                pass

    # 마감행(병합행) 건너뛰고 첫 번째 빈 행 찾기
    next_row = last_data_row + 1
    while next_row in merged_rows:
        next_row += 1
    next_no  = last_no + 1

    # 날짜 파싱
    from datetime import datetime
    try:
        y, m = period.split('-')
        import calendar
        last_day = calendar.monthrange(int(y), int(m))[1]
        start_date = datetime(int(y), int(m), 1)
        end_date   = datetime(int(y), int(m), last_day)
        period_str = f"{y}년 {int(m)}월"
    except:
        start_date = end_date = None
        period_str = period

    try:
        billing_dt = datetime.strptime(billing_date, '%Y-%m-%d') if billing_date else None
    except Exception:
        billing_dt = None

    if not rows:
        return jsonify({'error': '정산 데이터가 없습니다'}), 400

    for r in rows:
        sa    = r.get('supplyAmt', 0) or 0
        mr    = r.get('markupRate', 0) or 0
        af    = r.get('agencyFeeRate', 0) or 0
        markup    = round(sa * mr)
        agency_fee = round(sa * af)
        total     = sa + markup
        prev_diff  = r.get('prevDiff', 0) or 0
        billing_ad = sa + prev_diff
        billing_mk = markup
        billing_total = billing_ad + billing_mk
        diff = billing_ad - sa

        row_vals = {
            2:  next_no,
            3:  user_team,
            4:  user_name,
            5:  r.get('advertiser', adv_name),
            6:  r.get('campaign', camp_name),
            7:  r.get('media', ''),
            8:  r.get('accountId', ''),
            9:  period_str,
            10: start_date,
            11: end_date,
            12: sa,
            13: sa,           # 광고비 = 공급가액
            14: af if af else None,
            15: agency_fee if af else 0,
            16: mr if mr else None,
            17: markup if mr else 0,
            18: total,
            19: billing_dt,
            20: prev_diff if prev_diff else None,
            21: billing_ad,
            22: billing_mk if markup else None,
            23: billing_total,
            24: None,          # 마케팅프로그램비
            25: diff if diff else 0,
            26: r.get('note', ''),
            27: r.get('fxCurrency', 'KRW') or 'KRW',
            28: r.get('fxRate') if r.get('fxRate') else (1 if not r.get('fxCurrency') or r.get('fxCurrency') == 'KRW' else None),
        }

        import copy, re
        ref_row = last_data_row  # 마지막 실데이터 행 서식/수식 참조

        # 참조행에서 수식 패턴 읽기 (data_only=False로 열어야 수식 읽힘)
        # wb는 data_only=False로 열려있으므로 수식 그대로 읽힘
        ref_formulas = {}
        for ref_cell in ws[ref_row]:
            if ref_cell.column <= 33:
                v = ref_cell.value
                if isinstance(v, str) and v.startswith('='):
                    ref_formulas[ref_cell.column] = v

        for col, val in row_vals.items():
            cell = ws.cell(row=next_row, column=col)
            if isinstance(cell, MergedCell):
                continue

            # 참조행에 수식이 있는 열이면 → 수식 행번호만 바꿔서 적용
            if col in ref_formulas:
                formula = ref_formulas[col]
                # 수식 내 숫자(행번호)를 new_row로 교체 (알파벳 뒤 숫자만)
                new_formula = re.sub(
                    r'([A-Za-z]+)' + str(ref_row),
                    lambda m: m.group(1) + str(next_row),
                    formula
                )
                cell.value = new_formula
            else:
                cell.value = val

            # 참조 행 서식 복사
            ref_cell = ws.cell(row=ref_row, column=col)
            if not isinstance(ref_cell, MergedCell):
                try:
                    if ref_cell.has_style:
                        cell.font      = copy.copy(ref_cell.font)
                        cell.fill      = copy.copy(ref_cell.fill)
                        cell.border    = copy.copy(ref_cell.border)
                        cell.alignment = copy.copy(ref_cell.alignment)
                        cell.number_format = ref_cell.number_format
                except:
                    pass

        next_no  += 1
        next_row += 1
        # 다음 행도 병합행이면 건너뜀
        while next_row in merged_rows:
            next_row += 1

    # ── 월 마감 병합행 추가 ──
    try:
        # 기간에서 월 추출 (예: "2026-03" → "3월")
        month_label = f"{int(m)}월 마감"
        closing_row = next_row  # 마지막 데이터 다음 행

        # B:AB 병합
        ws.merge_cells(start_row=closing_row, start_column=2,
                       end_row=closing_row,   end_column=28)

        # B열에 "N월 마감" 텍스트
        closing_cell = ws.cell(row=closing_row, column=2, value=month_label)

        # 기존 마감행 서식 복사 (있으면)
        existing_closing = next((mr.min_row for mr in ws.merged_cells.ranges
                                  if mr.min_col == 2 and mr.max_col == 28
                                  and mr.min_row < closing_row), None)
        if existing_closing:
            ref_close = ws.cell(row=existing_closing, column=2)
            if ref_close.has_style:
                import copy
                closing_cell.font      = copy.copy(ref_close.font)
                closing_cell.fill      = copy.copy(ref_close.fill)
                closing_cell.border    = copy.copy(ref_close.border)
                closing_cell.alignment = copy.copy(ref_close.alignment)

        # merged_rows 갱신
        merged_rows.add(closing_row)
    except Exception as e:
        pass  # 마감행 추가 실패해도 나머지는 정상 저장

    # ── 세발요청 시트 처리 ──
    if '세발요청' in wb.sheetnames:
        ws2 = wb['세발요청']

        # 마지막 데이터 행 찾기
        last_row2 = 1
        for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row):
            if any(c.value is not None for c in row):
                last_row2 = row[0].row

        next_row2 = last_row2 + 1
        first = True

        for r in rows:
            billing_ad = (r.get('supplyAmt') or 0) + (r.get('prevDiff') or 0)
            billing_mk = round((r.get('supplyAmt') or 0) * (r.get('markupRate') or 0))

            ws2.cell(row=next_row2, column=1, value=billing_date if first else None)
            ws2.cell(row=next_row2, column=2, value=user_name if first else None)
            ws2.cell(row=next_row2, column=3, value='통합발행' if first else None)
            ws2.cell(row=next_row2, column=4, value=adv_biz_no if first else None)
            ws2.cell(row=next_row2, column=5, value=adv_name if first else None)
            ws2.cell(row=next_row2, column=6, value=adv_email if first else None)
            ws2.cell(row=next_row2, column=9, value='후불')
            ws2.cell(row=next_row2, column=10,
                value=f"{r.get('period','')} {r.get('campaign','')} {r.get('media','')} 매체운영비")
            c = ws2.cell(row=next_row2, column=11, value=billing_ad)
            c.number_format = '#,##0'
            ws2.cell(row=next_row2, column=14, value=r.get('note', ''))
            next_row2 += 1
            first = False

            if billing_mk > 0:
                ws2.cell(row=next_row2, column=9, value='후불')
                ws2.cell(row=next_row2, column=10,
                    value=f"{r.get('period','')} {r.get('campaign','')} {r.get('media','')} 대행수수료")
                c2 = ws2.cell(row=next_row2, column=11, value=billing_mk)
                c2.number_format = '#,##0'
                next_row2 += 1

    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"세금계산서발행요청파일_{adv_name}_{camp_name}_{period}.xlsx"
    return send_file(
        buf, as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

# ───────────── XLSX 내보내기 API ──────────────────────
@app.route('/api/export', methods=['POST'])
@require_login
def export_xlsx():
    d = request.json
    rows       = d.get('rows', [])
    adv_name   = d.get('advName', '')
    camp_name  = d.get('campName', '')
    period     = d.get('period', '')
    billing_date = d.get('billingDate', '')
    adv_biz_no = d.get('advBizNo', '')
    adv_email  = d.get('advEmail', '')
    user_name  = session.get('user_name', '')
    user_team  = session.get('user_team', '')

    wb = openpyxl.Workbook()

    # 헤더 스타일
    hdr_fill = PatternFill(start_color='D3D1C7', end_color='D3D1C7', fill_type='solid')
    hdr_font = Font(bold=True, size=10)
    hdr_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin = Side(style='thin', color='888780')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def style_header(ws, headers, col_widths):
        for col, (h, w) in enumerate(zip(headers, col_widths), 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = hdr_align
            cell.border = border
            ws.column_dimensions[get_column_letter(col)].width = w
        ws.row_dimensions[1].height = 30

    # ── 세일즈양식 시트 ──
    ws1 = wb.active
    ws1.title = '세일즈양식'
    ws1.freeze_panes = 'A2'

    h1 = ['번호','팀명','담당자','광고주명','캠페인명','매체명','계정ID',
          '매입월','소진시작','소진종료','공급가액(부가세별도)',
          '광고비','매체수수료율','매체수수료','마크업율','마크업수수료','합계',
          '청구일자','전월오차반영','광고비청구','마크업청구','청구합계','오차','비고','통화','환율']
    w1 = [5,8,10,20,14,12,16,10,12,12,16,14,10,12,8,14,16,14,12,14,12,14,12,22,6,6]
    style_header(ws1, h1, w1)

    num_cols = {11,12,14,16,17,19,20,21,22,23}  # 숫자 컬럼 (1-based)
    pct_cols = {13,15}

    for i, r in enumerate(rows, 2):
        vals = [
            r.get('no', i-1), user_team, user_name, r.get('advertiser',''),
            r.get('campaign',''), r.get('media',''), r.get('accountId',''),
            r.get('period',''), r.get('startDate',''), r.get('endDate',''),
            r.get('supplyAmt',0), r.get('supplyAmt',0),
            r.get('agencyFeeRate','') or '', r.get('agencyFee',0) or 0,
            r.get('markupRate','') or '', r.get('markup',0) or 0,
            r.get('total',0),
            r.get('billingDate',''), r.get('prevDiff',0),
            r.get('billingAdCost',0), r.get('billingMarkup',0),
            r.get('billingTotal',0), r.get('diff',0),
            r.get('note',''), r.get('fxCurrency','KRW') or 'KRW', r.get('fxRate') or 1
        ]
        for col, val in enumerate(vals, 1):
            cell = ws1.cell(row=i, column=col, value=val)
            cell.border = border
            if col in num_cols:
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal='right')
            elif col in pct_cols and isinstance(val, (int, float)):
                cell.number_format = '0.00%'
                cell.alignment = Alignment(horizontal='center')
            else:
                cell.alignment = Alignment(horizontal='left' if col > 4 else 'center')

    # ── 세발요청 시트 ──
    ws2 = wb.create_sheet('세발요청')
    h2 = ['작성일자','담당자명','발행종류','사업자번호','사업자명','이메일수신처','경영지원팀체크','발행담당자','선불/후불','항목명','공급가액','부가세','총액','비고']
    w2 = [14,10,10,14,20,24,14,10,8,36,16,10,16,24]
    style_header(ws2, h2, w2)

    row_num = 2
    first = True
    for r in rows:
        ws2.cell(row=row_num, column=1, value=billing_date if first else '')
        ws2.cell(row=row_num, column=2, value=user_name if first else '')
        ws2.cell(row=row_num, column=3, value='통합발행' if first else '')
        ws2.cell(row=row_num, column=4, value=adv_biz_no if first else '')
        ws2.cell(row=row_num, column=5, value=adv_name if first else '')
        ws2.cell(row=row_num, column=6, value=adv_email if first else '')
        ws2.cell(row=row_num, column=9, value='후불')
        ws2.cell(row=row_num, column=10, value=f"{r.get('period','')} {r.get('campaign','')} {r.get('media','')} 매체운영비")
        c = ws2.cell(row=row_num, column=11, value=r.get('billingAdCost', 0))
        c.number_format = '#,##0'
        ws2.cell(row=row_num, column=14, value=r.get('note', ''))
        for col in range(1, 15):
            ws2.cell(row=row_num, column=col).border = border
        row_num += 1
        first = False

        if (r.get('billingMarkup') or 0) > 0:
            ws2.cell(row=row_num, column=9, value='후불')
            ws2.cell(row=row_num, column=10, value=f"{r.get('period','')} {r.get('campaign','')} {r.get('media','')} 대행수수료")
            c2 = ws2.cell(row=row_num, column=11, value=r.get('billingMarkup', 0))
            c2.number_format = '#,##0'
            for col in range(1, 15):
                ws2.cell(row=row_num, column=col).border = border
            row_num += 1

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"정산_{adv_name}_{camp_name}_{period}.xlsx"
    return send_file(
        buf, as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


# ───────────── 프론트엔드 ─────────────────────────────
@app.route('/')
def index():
    return app.send_static_file('index.html')


if __name__ == '__main__':
    print("=" * 50)
    print("📊 광고 정산 자동화 시스템 시작")
    print("📌 접속 주소: http://localhost:5000")
    print("=" * 50)
    app.run(debug=True, port=5000)
