"""
광고 정산 자동화 시스템 - Flask 백엔드
실행: python app.py
접속: http://localhost:5000
"""

from flask import Flask, request, jsonify, session, send_file
import sqlite3, hashlib, os, io
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__, static_folder='static', static_url_path='')
app.secret_key = 'settlement_secret_key_change_in_production'

DB_PATH = 'settlement.db'


# ───────────── DB 초기화 ──────────────────────────────
def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    with get_db() as conn:
        conn.executescript('''
            CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username TEXT UNIQUE NOT NULL,
                password_hash TEXT NOT NULL,
                name TEXT NOT NULL,
                team TEXT DEFAULT ''
            );
            CREATE TABLE IF NOT EXISTS advertisers (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                name TEXT NOT NULL,
                biz_no TEXT DEFAULT '',
                email TEXT DEFAULT '',
                contact_name TEXT DEFAULT ''
            );
            CREATE TABLE IF NOT EXISTS campaigns (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                advertiser_id INTEGER NOT NULL,
                name TEXT NOT NULL,
                sort_order INTEGER DEFAULT 0
            );
            CREATE TABLE IF NOT EXISTS media_rates (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                campaign_id INTEGER NOT NULL,
                media TEXT NOT NULL,
                markup_rate REAL DEFAULT 0,
                agency_fee_rate REAL DEFAULT 0,
                payback_rate REAL DEFAULT 0,
                sort_order INTEGER DEFAULT 0
            );
            CREATE TABLE IF NOT EXISTS settlements (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                advertiser TEXT, campaign TEXT, media TEXT,
                period TEXT, start_date TEXT, end_date TEXT,
                supply_amt REAL DEFAULT 0,
                markup_rate REAL DEFAULT 0, markup REAL DEFAULT 0,
                agency_fee_rate REAL DEFAULT 0, agency_fee REAL DEFAULT 0,
                total REAL DEFAULT 0,
                billing_date TEXT,
                prev_diff REAL DEFAULT 0,
                billing_ad_cost REAL DEFAULT 0,
                billing_markup REAL DEFAULT 0,
                billing_total REAL DEFAULT 0,
                diff REAL DEFAULT 0,
                account_id TEXT DEFAULT '',
                note TEXT DEFAULT '',
                created_at TEXT
            );
        ''')

init_db()


# ───────────── 유틸 ───────────────────────────────────
def hash_pw(pw):
    return hashlib.sha256(pw.encode('utf-8')).hexdigest()

def require_login(f):
    from functools import wraps
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return jsonify({'error': '로그인이 필요합니다'}), 401
        return f(*args, **kwargs)
    return decorated


# ───────────── 인증 API ───────────────────────────────
@app.route('/api/register', methods=['POST'])
def register():
    d = request.json
    if not d.get('username') or not d.get('password') or not d.get('name'):
        return jsonify({'error': '필수 항목을 입력해주세요'}), 400
    try:
        with get_db() as conn:
            conn.execute(
                'INSERT INTO users (username, password_hash, name, team) VALUES (?,?,?,?)',
                (d['username'], hash_pw(d['password']), d['name'], d.get('team', ''))
            )
        return jsonify({'success': True})
    except sqlite3.IntegrityError:
        return jsonify({'error': '이미 사용 중인 아이디입니다'}), 400

@app.route('/api/login', methods=['POST'])
def login():
    d = request.json
    with get_db() as conn:
        user = conn.execute(
            'SELECT * FROM users WHERE username=? AND password_hash=?',
            (d.get('username', ''), hash_pw(d.get('password', '')))
        ).fetchone()
    if not user:
        return jsonify({'error': '아이디 또는 비밀번호가 틀렸습니다'}), 401
    session.permanent = True
    session['user_id'] = user['id']
    session['user_name'] = user['name']
    session['user_team'] = user['team']
    return jsonify({'success': True, 'name': user['name'], 'team': user['team']})

@app.route('/api/logout', methods=['POST'])
def logout():
    session.clear()
    return jsonify({'success': True})

@app.route('/api/me')
def me():
    if 'user_id' not in session:
        return jsonify({'error': 'Unauthorized'}), 401
    return jsonify({
        'id': session['user_id'],
        'name': session['user_name'],
        'team': session['user_team']
    })


# ───────────── 광고주 API ─────────────────────────────
def get_adv_full(conn, adv_id):
    adv = conn.execute('SELECT * FROM advertisers WHERE id=?', (adv_id,)).fetchone()
    if not adv:
        return None
    camps = conn.execute('SELECT * FROM campaigns WHERE advertiser_id=? ORDER BY sort_order', (adv_id,)).fetchall()
    campaigns = []
    for c in camps:
        rates = conn.execute('SELECT * FROM media_rates WHERE campaign_id=? ORDER BY sort_order', (c['id'],)).fetchall()
        campaigns.append({
            'id': c['id'], 'name': c['name'],
            'mediaRates': [{'id': r['id'], 'media': r['media'],
                            'markupRate': r['markup_rate'],
                            'agencyFeeRate': r['agency_fee_rate'],
                            'paybackRate': r['payback_rate']} for r in rates]
        })
    return {
        'id': adv['id'], 'name': adv['name'],
        'bizNo': adv['biz_no'], 'email': adv['email'],
        'contactName': adv['contact_name'], 'campaigns': campaigns
    }

@app.route('/api/advertisers')
@require_login
def get_advertisers():
    with get_db() as conn:
        advs = conn.execute('SELECT id FROM advertisers WHERE user_id=?', (session['user_id'],)).fetchall()
        return jsonify([get_adv_full(conn, a['id']) for a in advs])

@app.route('/api/advertisers', methods=['POST'])
@require_login
def create_advertiser():
    d = request.json
    with get_db() as conn:
        cur = conn.execute(
            'INSERT INTO advertisers (user_id, name, biz_no, email, contact_name) VALUES (?,?,?,?,?)',
            (session['user_id'], d['name'], d.get('bizNo',''), d.get('email',''), d.get('contactName',''))
        )
        adv_id = cur.lastrowid
        for ci, camp in enumerate(d.get('campaigns', [])):
            cur2 = conn.execute('INSERT INTO campaigns (advertiser_id, name, sort_order) VALUES (?,?,?)', (adv_id, camp['name'], ci))
            camp_id = cur2.lastrowid
            for mi, mr in enumerate(camp.get('mediaRates', [])):
                conn.execute(
                    'INSERT INTO media_rates (campaign_id, media, markup_rate, agency_fee_rate, payback_rate, sort_order) VALUES (?,?,?,?,?,?)',
                    (camp_id, mr['media'], mr.get('markupRate',0), mr.get('agencyFeeRate',0), mr.get('paybackRate',0), mi)
                )
    return jsonify({'success': True, 'id': adv_id})

@app.route('/api/advertisers/<int:adv_id>', methods=['PUT'])
@require_login
def update_advertiser(adv_id):
    d = request.json
    with get_db() as conn:
        conn.execute(
            'UPDATE advertisers SET name=?, biz_no=?, email=?, contact_name=? WHERE id=? AND user_id=?',
            (d['name'], d.get('bizNo',''), d.get('email',''), d.get('contactName',''), adv_id, session['user_id'])
        )
        camps = conn.execute('SELECT id FROM campaigns WHERE advertiser_id=?', (adv_id,)).fetchall()
        for c in camps:
            conn.execute('DELETE FROM media_rates WHERE campaign_id=?', (c['id'],))
        conn.execute('DELETE FROM campaigns WHERE advertiser_id=?', (adv_id,))
        for ci, camp in enumerate(d.get('campaigns', [])):
            cur = conn.execute('INSERT INTO campaigns (advertiser_id, name, sort_order) VALUES (?,?,?)', (adv_id, camp['name'], ci))
            camp_id = cur.lastrowid
            for mi, mr in enumerate(camp.get('mediaRates', [])):
                conn.execute(
                    'INSERT INTO media_rates (campaign_id, media, markup_rate, agency_fee_rate, payback_rate, sort_order) VALUES (?,?,?,?,?,?)',
                    (camp_id, mr['media'], mr.get('markupRate',0), mr.get('agencyFeeRate',0), mr.get('paybackRate',0), mi)
                )
    return jsonify({'success': True})

@app.route('/api/advertisers/<int:adv_id>', methods=['DELETE'])
@require_login
def delete_advertiser(adv_id):
    with get_db() as conn:
        camps = conn.execute('SELECT id FROM campaigns WHERE advertiser_id=?', (adv_id,)).fetchall()
        for c in camps:
            conn.execute('DELETE FROM media_rates WHERE campaign_id=?', (c['id'],))
        conn.execute('DELETE FROM campaigns WHERE advertiser_id=?', (adv_id,))
        conn.execute('DELETE FROM advertisers WHERE id=? AND user_id=?', (adv_id, session['user_id']))
    return jsonify({'success': True})


# ───────────── 정산 API ───────────────────────────────
@app.route('/api/settlements')
@require_login
def get_settlements():
    with get_db() as conn:
        rows = conn.execute(
            'SELECT * FROM settlements WHERE user_id=? ORDER BY created_at DESC LIMIT 200',
            (session['user_id'],)
        ).fetchall()
    return jsonify([dict(r) for r in rows])

@app.route('/api/settlements', methods=['POST'])
@require_login
def save_settlements():
    rows = request.json
    now = datetime.now().isoformat()
    with get_db() as conn:
        for r in rows:
            conn.execute('''
                INSERT INTO settlements (user_id, advertiser, campaign, media, period,
                    start_date, end_date, supply_amt, markup_rate, markup,
                    agency_fee_rate, agency_fee, total, billing_date, prev_diff,
                    billing_ad_cost, billing_markup, billing_total, diff,
                    account_id, note, created_at)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            ''', (
                session['user_id'], r.get('advertiser'), r.get('campaign'), r.get('media'),
                r.get('period'), r.get('startDate'), r.get('endDate'),
                r.get('supplyAmt',0), r.get('markupRate',0), r.get('markup',0),
                r.get('agencyFeeRate',0), r.get('agencyFee',0), r.get('total',0),
                r.get('billingDate'), r.get('prevDiff',0), r.get('billingAdCost',0),
                r.get('billingMarkup',0), r.get('billingTotal',0), r.get('diff',0),
                r.get('accountId',''), r.get('note',''), now
            ))
    return jsonify({'success': True})


# ───────────── 파일 파싱 API ──────────────────────────
@app.route('/api/parse/naver', methods=['POST'])
@require_login
def parse_naver():
    if 'file' not in request.files:
        return jsonify({'error': '파일이 없습니다'}), 400

    file = request.files['file']
    ext = file.filename.rsplit('.', 1)[-1].lower() if '.' in file.filename else ''
    result = {}

    try:
        if ext == 'xls':
            import xlrd
            wb = xlrd.open_workbook(file_contents=file.read())
            sheet_name = next((n for n in wb.sheet_names() if '세금' in n or '계산서' in n), None)
            sheet = wb.sheet_by_name(sheet_name) if sheet_name else wb.sheet_by_index(0)

            hdr_row, cols = -1, {}
            for i in range(min(sheet.nrows, 10)):
                vals = [str(sheet.cell_value(i, j)) for j in range(sheet.ncols)]
                if '계정명' in vals or '공급가액' in vals:
                    hdr_row = i
                    cols = {v.strip(): j for j, v in enumerate(vals)}
                    break

            if hdr_row >= 0:
                for i in range(hdr_row + 1, sheet.nrows):
                    acct_col = cols.get('계정명', -1)
                    amt_col  = cols.get('공급가액', -1)
                    if acct_col < 0 or amt_col < 0:
                        continue
                    acct = str(sheet.cell_value(i, acct_col)).strip()
                    try:
                        amt = float(sheet.cell_value(i, amt_col))
                        if acct and amt:
                            result[acct] = result.get(acct, 0) + amt
                    except (ValueError, TypeError):
                        pass

        else:  # xlsx
            wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
            sheet_name = next((n for n in wb.sheetnames if '세금' in n or '계산서' in n), None)
            ws = wb[sheet_name] if sheet_name else wb.active

            rows_data = list(ws.iter_rows(values_only=True))
            hdr_row, cols = -1, {}
            for i, row in enumerate(rows_data[:10]):
                if not row:
                    continue
                strs = [str(c).strip() if c is not None else '' for c in row]
                if '계정명' in strs or '공급가액' in strs:
                    hdr_row = i
                    cols = {v: j for j, v in enumerate(strs)}
                    break

            if hdr_row >= 0:
                for row in rows_data[hdr_row + 1:]:
                    if not row:
                        continue
                    acct_col = cols.get('계정명', -1)
                    amt_col  = cols.get('공급가액', -1)
                    if acct_col < 0 or amt_col < 0:
                        continue
                    acct = str(row[acct_col]).strip() if row[acct_col] is not None else ''
                    amt  = row[amt_col]
                    if acct and acct not in ('None', '') and isinstance(amt, (int, float)) and amt:
                        result[acct] = result.get(acct, 0) + float(amt)

    except Exception as e:
        return jsonify({'error': f'파싱 오류: {str(e)}'}), 500

    return jsonify({'data': result, 'total': sum(result.values())})


# ───────────── XLSX 내보내기 API ──────────────────────
@app.route('/api/export', methods=['POST'])
@require_login
def export_xlsx():
    d = request.json
    rows       = d.get('rows', [])
    adv_name   = d.get('advName', '')
    camp_name  = d.get('campName', '')
    period     = d.get('period', '')
    billing_date = d.get('billingDate', '')
    adv_biz_no = d.get('advBizNo', '')
    adv_email  = d.get('advEmail', '')
    user_name  = session.get('user_name', '')
    user_team  = session.get('user_team', '')

    wb = openpyxl.Workbook()

    # 헤더 스타일
    hdr_fill = PatternFill(start_color='D3D1C7', end_color='D3D1C7', fill_type='solid')
    hdr_font = Font(bold=True, size=10)
    hdr_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin = Side(style='thin', color='888780')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def style_header(ws, headers, col_widths):
        for col, (h, w) in enumerate(zip(headers, col_widths), 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = hdr_align
            cell.border = border
            ws.column_dimensions[get_column_letter(col)].width = w
        ws.row_dimensions[1].height = 30

    # ── 세일즈양식 시트 ──
    ws1 = wb.active
    ws1.title = '세일즈양식'
    ws1.freeze_panes = 'A2'

    h1 = ['번호','팀명','담당자','광고주명','캠페인명','매체명','계정ID',
          '매입월','소진시작','소진종료','공급가액(부가세별도)',
          '광고비','매체수수료율','매체수수료','마크업율','마크업수수료','합계',
          '청구일자','전월오차반영','광고비청구','마크업청구','청구합계','오차','비고','통화','환율']
    w1 = [5,8,10,20,14,12,16,10,12,12,16,14,10,12,8,14,16,14,12,14,12,14,12,22,6,6]
    style_header(ws1, h1, w1)

    num_cols = {11,12,14,16,17,19,20,21,22,23}  # 숫자 컬럼 (1-based)
    pct_cols = {13,15}

    for i, r in enumerate(rows, 2):
        vals = [
            r.get('no', i-1), user_team, user_name, r.get('advertiser',''),
            r.get('campaign',''), r.get('media',''), r.get('accountId',''),
            r.get('period',''), r.get('startDate',''), r.get('endDate',''),
            r.get('supplyAmt',0), r.get('supplyAmt',0),
            r.get('agencyFeeRate','') or '', r.get('agencyFee',0) or 0,
            r.get('markupRate','') or '', r.get('markup',0) or 0,
            r.get('total',0),
            r.get('billingDate',''), r.get('prevDiff',0),
            r.get('billingAdCost',0), r.get('billingMarkup',0),
            r.get('billingTotal',0), r.get('diff',0),
            r.get('note',''), 'KRW', 1
        ]
        for col, val in enumerate(vals, 1):
            cell = ws1.cell(row=i, column=col, value=val)
            cell.border = border
            if col in num_cols:
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal='right')
            elif col in pct_cols and isinstance(val, (int, float)):
                cell.number_format = '0.00%'
                cell.alignment = Alignment(horizontal='center')
            else:
                cell.alignment = Alignment(horizontal='left' if col > 4 else 'center')

    # ── 세발요청 시트 ──
    ws2 = wb.create_sheet('세발요청')
    h2 = ['작성일자','담당자명','발행종류','사업자번호','사업자명','이메일수신처','경영지원팀체크','발행담당자','선불/후불','항목명','공급가액','부가세','총액','비고']
    w2 = [14,10,10,14,20,24,14,10,8,36,16,10,16,24]
    style_header(ws2, h2, w2)

    row_num = 2
    first = True
    for r in rows:
        ws2.cell(row=row_num, column=1, value=billing_date if first else '')
        ws2.cell(row=row_num, column=2, value=user_name if first else '')
        ws2.cell(row=row_num, column=3, value='통합발행' if first else '')
        ws2.cell(row=row_num, column=4, value=adv_biz_no if first else '')
        ws2.cell(row=row_num, column=5, value=adv_name if first else '')
        ws2.cell(row=row_num, column=6, value=adv_email if first else '')
        ws2.cell(row=row_num, column=9, value='후불')
        ws2.cell(row=row_num, column=10, value=f"{r.get('period','')} {r.get('campaign','')} {r.get('media','')} 매체운영비")
        c = ws2.cell(row=row_num, column=11, value=r.get('billingAdCost', 0))
        c.number_format = '#,##0'
        ws2.cell(row=row_num, column=14, value=r.get('note', ''))
        for col in range(1, 15):
            ws2.cell(row=row_num, column=col).border = border
        row_num += 1
        first = False

        if (r.get('billingMarkup') or 0) > 0:
            ws2.cell(row=row_num, column=9, value='후불')
            ws2.cell(row=row_num, column=10, value=f"{r.get('period','')} {r.get('campaign','')} {r.get('media','')} 대행수수료")
            c2 = ws2.cell(row=row_num, column=11, value=r.get('billingMarkup', 0))
            c2.number_format = '#,##0'
            for col in range(1, 15):
                ws2.cell(row=row_num, column=col).border = border
            row_num += 1

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"정산_{adv_name}_{camp_name}_{period}.xlsx"
    return send_file(
        buf, as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


# ───────────── 프론트엔드 ─────────────────────────────
@app.route('/')
def index():
    return app.send_static_file('index.html')


if __name__ == '__main__':
    print("=" * 50)
    print("📊 광고 정산 자동화 시스템 시작")
    print("📌 접속 주소: http://localhost:5000")
    print("=" * 50)
    app.run(debug=True, port=5000)
